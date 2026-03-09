"""
Walk-Forward Backtesting Script

Trains on a rolling 2-year window, predicts on the next year, slides forward
through all available data. No pre-trained models required — this script trains
and tests every window from scratch (with warm-starting between adjacent windows).

Example: data available 2018–2026
  Window 1:  Train 2018–2019  →  Test 2020
  Window 2:  Train 2019–2020  →  Test 2021
  Window 3:  Train 2020–2021  →  Test 2022
  Window 4:  Train 2021–2022  →  Test 2023
  Window 5:  Train 2022–2023  →  Test 2024
  Window 6:  Train 2023–2024  →  Test 2025
  Window 7:  Train 2024–2025  →  Test 2026

Odds
----
All P&L calculations use Betfair Lay odds (Lay_odds, Lay_top5, Lay_top10,
Lay_top20), which are pre-event snapshots and closer to expected back prices
than bookie odds. No pre-computed profit columns are used; P&L is formula-based
for all markets with dead-heat adjustment applied to place markets.

Dead heat (place markets only)
-------------------------------
A dead heat occurs when players tie at exactly the cut position (5, 10, or 20),
meaning fewer places are available than players tied there. Only players AT the
cut position are affected; players finishing clearly inside the cut get full odds.

  places_filled_above     = count of players with posn < cut
  places_available_in_tie = cut - places_filled_above
  players_tied            = count of players with posn == cut position in data
  RF                      = places_available_in_tie / players_tied  (clamped 0–1)

  Win P&L  = stake × (RF × lay_odds - 1)
  Loss P&L = -stake

Caching
-------
Trained model bundles are saved to Output/WalkForward/Models/{tour}_{year}/
so the script can be interrupted and resumed without retraining completed windows.
Use --force-retrain to ignore the cache and retrain everything.

Optuna trials
-------------
Default is 30 trials per model per market (vs 75 in production) to keep runtimes
manageable across many windows. Use --trials 75 for full production quality.
Warm-start: best params from window N are passed as starting suggestions to window N+1.

Back betting
------------
Bet when Normalised_Model_Odds < Lay_odds (fixed £10 stake).
Winner:          P&L = (lay_odds - 1) × 10 on win, -10 on loss.
Top5/Top10/Top20: same formula with dead-heat RF applied to winning rows.

Grid search
-----------
After all windows are processed a parameter grid is swept over the combined
All_Predictions data. Dimensions:
  Edge threshold  — ratio of lay odds to model odds required before betting
  Min/Max odds    — odds range filter
  Min rating      — player rating floor filter
Results are written to the Strategy_Grid sheet of the output workbook.

Output
------
Output/WalkForward/Results/{tour}_WalkForward_Backtest.xlsx
  Summary         — overall metrics + back P&L across all windows, per market
  Season_Summary  — same broken down by test year
  Event_Results   — per-event P&L with cumulative P&L column
  All_Predictions — every player prediction + outcome + bet P&L
  Calib_*         — calibration bins (predicted vs actual) per market
  Strategy_Grid   — grid search results sorted by ROI

Run:
  python walk_forward_backtest.py                         # both tours, all windows, 30 trials
  python walk_forward_backtest.py --tour PGA
  python walk_forward_backtest.py --trials 75             # full Optuna
  python walk_forward_backtest.py --min-year 2022         # start from 2022 test window
  python walk_forward_backtest.py --force-retrain         # ignore cached models
"""

import argparse
import shutil
import sys
import warnings
from itertools import product
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import (
    average_precision_score,
    brier_score_loss,
    log_loss,
    roc_auc_score,
)

try:
    sys.path.insert(0, str(Path(__file__).parent))
except NameError:
    _cwd = Path.cwd()
    _candidates = [_cwd, _cwd / "Weekly_Modelling_Python" / "Script", _cwd / "Script"]
    _script_dir = next((p for p in _candidates if (p / "seasonal_model_training.py").exists()), _cwd)
    sys.path.insert(0, str(_script_dir))

import seasonal_model_training as smt

from config import (
    BASE_DIR,
    BETTING_MARKETS,
    PREDICTIONS_DIR,
    TOUR_CONFIG,
    TRAINING_YEARS,
)
from seasonal_model_training import get_market_vars, tss_optimal

warnings.filterwarnings("ignore")

# ===== PATHS =====
WF_DIR         = BASE_DIR / "Output" / "WalkForward"
WF_MODELS_DIR  = WF_DIR / "Models"
WF_RESULTS_DIR = WF_DIR / "Results"

# ===== CONSTANTS =====
BACK_STAKE        = 10.0   # £ per back bet
DEFAULT_WF_TRIALS = 30
MIN_POSITIVES     = 10     # skip a market window if training set has fewer positives

# ===== LAY ODDS COLUMN MAPPING =====
# Maps each market name to its Betfair Lay odds column.
# These are pre-event snapshots used for both the edge condition and P&L.
LAY_ODDS_COLS = {
    "Winner": "Lay_odds",
    "Top5":   "Lay_top5",
    "Top10":  "Lay_top10",
    "Top20":  "Lay_top20",
}

# Cut position for each place market — used for dead-heat calculation.
# Winner has no cut position (no dead-heat logic applied).
PLACE_MARKET_CUTS = {
    "Top5":  5,
    "Top10": 10,
    "Top20": 20,
}

# ===== GRID SEARCH PARAMETERS =====
GRID_EDGE_THRESHOLDS = [1.0, 1.05, 1.1, 1.25, 1.5, 2.0, 3.0, 5.0]
GRID_MIN_ODDS        = [1.0, 3.0, 5.0, 10.0]
GRID_MAX_ODDS        = [9999, 10, 25, 50, 100, 200, 500]   # 9999 = no cap
GRID_MIN_RATING      = [None, 50, 55, 60, 65, 70]           # None = no filter


# ===== CLI =====

def parse_args():
    parser = argparse.ArgumentParser(
        description="Walk-forward backtest for golf prediction models"
    )
    parser.add_argument("--tour", default=None,
                        help="Tour to backtest: PGA or Euro (default: both)")
    parser.add_argument("--trials", type=int, default=DEFAULT_WF_TRIALS,
                        help=f"Optuna trials per model per market "
                             f"(default: {DEFAULT_WF_TRIALS})")
    parser.add_argument("--min-year", type=int, default=None, dest="min_year",
                        help="Earliest test year to evaluate (default: all available)")
    parser.add_argument("--force-retrain", action="store_true", dest="force_retrain",
                        help="Ignore cached models and retrain all windows")
    return parser.parse_args()


# ===== PREDICTION =====

def predict_event(event_df: pd.DataFrame, market_name: str,
                  market_pkg: dict) -> pd.DataFrame | None:
    """
    Apply a trained market package to a single event's player rows.
    Mirrors the weekly prediction script exactly so backtest reflects
    real deployment behaviour.
    """
    model_vars = market_pkg["model_vars"]
    odds_col   = market_pkg["odds_col"]

    available = [v for v in model_vars if v in event_df.columns]
    df = event_df[available + [odds_col]].copy()
    for col in available + [odds_col]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna()
    if len(df) == 0:
        return None

    X    = df[available].values.astype(float)
    odds = df[odds_col].values.astype(float)

    model_preds = np.column_stack([
        m.predict_proba(X)[:, 1] for m in market_pkg["models"].values()
    ])
    raw_score = model_preds.mean(axis=1)

    meta_X_scaled = market_pkg["meta_scaler"].transform(model_preds)
    proba         = market_pkg["meta_model"].predict_proba(meta_X_scaled)[:, 1]

    market_size = market_pkg["market_size"]
    prob_sum    = proba.sum()
    norm_prob   = (proba / prob_sum) * market_size if prob_sum > 0 else proba

    result = event_df.loc[df.index].copy()
    result["Model_Score"]            = raw_score.round(5)
    result["Probability"]            = proba.round(6)
    result["Normalised_Probability"] = norm_prob.round(6)
    result["Normalised_Model_Odds"]  = (1.0 / norm_prob.clip(1e-8)).round(2)
    return result


# ===== DEAD HEAT =====

def compute_reduction_factors(event_df: pd.DataFrame, cut: int) -> pd.Series:
    """
    Compute a per-row dead-heat reduction factor (RF) for a place market.

    Only rows where posn == the tied cut position receive RF < 1.
    Rows finishing clearly inside the cut (posn < cut) always get RF = 1.0
    because they are unambiguous winners regardless of any tie at the boundary.

    Parameters
    ----------
    event_df : DataFrame for a single event (all players).
    cut      : Cut position for the market (5, 10, or 20).

    Returns
    -------
    pd.Series of RF values aligned to event_df.index.
    """
    posn = pd.to_numeric(event_df["posn"], errors="coerce")
    rf   = pd.Series(1.0, index=event_df.index)

    # Count players who finished clearly inside the cut
    places_filled_above = int((posn < cut).sum())

    # How many places remain to be filled by the tied group at the cut
    places_available = cut - places_filled_above

    # Players tied exactly at the cut position in the data
    # (posn holds the tied value, e.g. 4 when four players tie for 4th)
    tied_mask    = posn == cut
    players_tied = int(tied_mask.sum())

    if players_tied > 1 and places_available < players_tied:
        # Genuine dead heat — clamp RF to [0, 1] as a safety guard
        raw_rf = places_available / players_tied
        dh_rf  = float(np.clip(raw_rf, 0.0, 1.0))
        rf[tied_mask] = dh_rf

    return rf


# ===== BACK BETTING STRATEGY =====

def apply_back_strategy(df: pd.DataFrame, lay_odds_col: str,
                        target_col: str,
                        market_name: str = "Winner") -> pd.DataFrame:
    """
    Back when Normalised_Model_Odds < Lay_odds (model thinks player underpriced).
    Fixed stake BACK_STAKE per bet.

    Winner market — straight formula, no dead-heat adjustment:
      Win:  P&L = (lay_odds - 1) * BACK_STAKE
      Loss: P&L = -BACK_STAKE

    Place markets (Top5 / Top10 / Top20) — dead-heat RF applied to winners
    who are tied exactly at the cut position:
      Win (clear):       P&L = (lay_odds - 1) * BACK_STAKE          [RF = 1]
      Win (dead heat):   P&L = BACK_STAKE * (RF * lay_odds - 1)     [RF < 1]
      Loss:              P&L = -BACK_STAKE

    The RF is computed per-event from the posn column. If posn is missing
    the function falls back to RF = 1.0 (no adjustment) for that event.
    """
    df     = df.copy()
    lay    = df[lay_odds_col].to_numpy(dtype=float)
    actual = df[target_col].to_numpy(dtype=float)
    model  = df["Normalised_Model_Odds"].to_numpy(dtype=float)
    is_back = model < lay

    # Build RF array — default 1.0 (no dead heat)
    rf = np.ones(len(df), dtype=float)

    cut = PLACE_MARKET_CUTS.get(market_name)
    if cut is not None and "posn" in df.columns:
        # Compute RF per event so ties are evaluated within each field
        for event_id, idx in df.groupby("eventID").groups.items():
            event_slice = df.loc[idx]
            event_rf    = compute_reduction_factors(event_slice, cut)
            rf[df.index.get_indexer(idx)] = event_rf.values

    # P&L: stake × (RF × odds - 1) on win, -stake on loss
    # When RF == 1 this reduces to the standard (odds - 1) × stake formula
    back_pnl = np.where(
        is_back & (actual == 1),  BACK_STAKE * (rf * lay - 1),
        np.where(is_back, -BACK_STAKE, 0.0)
    )

    df["Back_Bet"]          = is_back
    df["Back_PnL"]          = back_pnl
    df["DeadHeat_RF"]       = rf          # retained for inspection / audit
    return df


# ===== STRATEGY SUMMARY =====

def back_summary(df: pd.DataFrame, target_col: str) -> dict:
    bets = df[df["Back_Bet"]]
    if len(bets) == 0:
        return {"Back_NBets": 0, "Back_NWon": 0, "Back_HitRate": np.nan,
                "Back_PnL": 0.0, "Back_ROI": np.nan}
    n_bets = len(bets)
    n_won  = int(bets[target_col].sum())
    pnl    = float(bets["Back_PnL"].sum())
    return {
        "Back_NBets":   n_bets,
        "Back_NWon":    n_won,
        "Back_HitRate": round(n_won / n_bets, 4),
        "Back_PnL":     round(pnl, 2),
        "Back_ROI":     round(pnl / (n_bets * BACK_STAKE), 4),
    }


# ===== MODEL QUALITY METRICS =====

def compute_discrimination(y_true: np.ndarray, y_prob: np.ndarray) -> dict:
    if y_true.sum() == 0:
        return {"AUC": np.nan, "Avg_Precision": np.nan, "TSS": np.nan,
                "Log_Loss": np.nan, "Brier": np.nan}
    y_clip = np.clip(y_prob, 1e-15, 1 - 1e-15)
    return {
        "AUC":           round(roc_auc_score(y_true, y_prob), 4),
        "Avg_Precision": round(average_precision_score(y_true, y_prob), 4),
        "TSS":           round(tss_optimal(y_true, y_prob), 4),
        "Log_Loss":      round(log_loss(y_true, y_clip), 5),
        "Brier":         round(brier_score_loss(y_true, y_prob), 5),
    }


def calibration_bins(y_true: np.ndarray, y_prob: np.ndarray,
                     n_bins: int = 10) -> pd.DataFrame:
    bins = np.linspace(0, 1, n_bins + 1)
    rows = []
    for lo, hi in zip(bins[:-1], bins[1:]):
        mask = (y_prob >= lo) & (y_prob < hi)
        if mask.sum() == 0:
            continue
        rows.append({
            "Bin_Low":        round(lo, 2),
            "Bin_High":       round(hi, 2),
            "N":              int(mask.sum()),
            "Mean_Predicted": round(float(y_prob[mask].mean()), 4),
            "Actual_Rate":    round(float(y_true[mask].mean()), 4),
        })
    return pd.DataFrame(rows)


# ===== WALK-FORWARD WINDOW LOGIC =====

def get_windows(df: pd.DataFrame, min_test_year: int = None):
    """
    Build list of (train_start, train_end, test_year) tuples.
    Requires at least TRAINING_YEARS of prior data for each test year.
    """
    years = sorted(df["Date"].dt.year.dropna().unique())
    windows = []
    for test_year in years:
        train_end   = test_year - 1
        train_start = test_year - TRAINING_YEARS
        if train_start < years[0]:
            continue
        if min_test_year and test_year < min_test_year:
            continue
        windows.append((int(train_start), int(train_end), int(test_year)))
    return windows


# ===== WINDOW TRAINING =====

def train_window(tour_key: str, train_df: pd.DataFrame, test_year: int,
                 n_trials: int, prev_window_dir: Path = None) -> dict:
    """
    Train all four betting markets for one walk-forward window.

    Monkey-patches smt.OPTUNA_TRIALS and smt.MODELS_DIR so tuning uses
    the walk-forward trial budget and saves/loads warm params to the
    window-specific cache directory rather than the production MODELS_DIR.

    Warm-start: best params from prev_window_dir are copied into this
    window's directory before training begins.

    Returns dict: market_name → market_package (or empty dict if no markets trained).
    """
    window_dir = WF_MODELS_DIR / f"{tour_key}_{test_year}"
    window_dir.mkdir(parents=True, exist_ok=True)

    if prev_window_dir and prev_window_dir.exists():
        for f in prev_window_dir.glob("*_best_params.pkl"):
            dest = window_dir / f.name
            if not dest.exists():
                shutil.copy(f, dest)

    orig_trials     = smt.OPTUNA_TRIALS
    orig_models_dir = smt.MODELS_DIR
    smt.OPTUNA_TRIALS = n_trials
    smt.MODELS_DIR    = window_dir

    try:
        package = {}
        for market_name, market_config in BETTING_MARKETS.items():
            target_col = market_config["target_col"]

            if target_col in train_df.columns:
                n_pos = int(train_df[target_col].sum())
                if n_pos < MIN_POSITIVES:
                    print(f"    {market_name}: skipped "
                          f"({n_pos} positives < {MIN_POSITIVES} minimum)")
                    continue

            model_vars = get_market_vars(market_config)
            result = smt.train_market(
                market_name, market_config, train_df, tour_key, model_vars
            )
            if result is not None:
                package[market_name] = result
    finally:
        smt.OPTUNA_TRIALS = orig_trials
        smt.MODELS_DIR    = orig_models_dir

    return package


# ===== WINDOW BACKTESTING =====

def backtest_window(tour_key: str, test_df: pd.DataFrame, test_year: int,
                    package: dict):
    """
    Apply all trained markets to every event in test_df using Betfair Lay odds.
    Returns (all_predictions list, event_summaries list).
    """
    all_predictions = []
    event_summaries = []

    events = test_df["eventID"].unique()
    print(f"  Test {test_year}: {len(events)} events  |  {len(test_df):,} player-rows")

    for event_id in sorted(events):
        event_df   = test_df[test_df["eventID"] == event_id].copy()
        event_date = event_df["Date"].iloc[0].strftime("%Y-%m-%d")

        for market_name, market_pkg in package.items():
            market_config = BETTING_MARKETS[market_name]
            target_col    = market_config["target_col"]
            odds_col      = market_config["odds_col"]
            lay_odds_col  = LAY_ODDS_COLS.get(market_name)

            if target_col not in event_df.columns:
                continue

            # Skip if Lay odds column is absent or fully missing for this event
            if lay_odds_col is None or lay_odds_col not in event_df.columns:
                print(f"    {market_name}: no Lay odds column '{lay_odds_col}' — skipping event {event_id}")
                continue
            if event_df[lay_odds_col].isna().all():
                continue

            preds = predict_event(event_df, market_name, market_pkg)
            if preds is None or len(preds) == 0:
                continue

            preds["Actual"] = preds[target_col].astype(int)
            preds = apply_back_strategy(preds, lay_odds_col, target_col, market_name)

            bs = back_summary(preds, target_col)

            event_summaries.append({
                "Tour":      tour_key,
                "Test_Year": test_year,
                "EventID":   event_id,
                "Date":      event_date,
                "Market":    market_name,
                "FieldSize": len(preds),
                "Positives": int(preds["Actual"].sum()),
                **bs,
            })

            preds["Tour"]      = tour_key
            preds["Test_Year"] = test_year
            preds["EventID"]   = event_id
            preds["Date"]      = event_date
            preds["Market"]    = market_name
            all_predictions.append(preds)

    return all_predictions, event_summaries


# ===== AGGREGATE ACROSS WINDOWS =====

def aggregate_results(all_preds_list: list, event_summaries_list: list,
                      tour_key: str) -> dict:
    """Combine results from all walk-forward windows into summary DataFrames."""
    if not all_preds_list:
        return None

    pred_df  = pd.concat(all_preds_list, ignore_index=True)
    event_df = (
        pd.DataFrame(event_summaries_list)
        .sort_values(["Market", "Test_Year", "Date", "EventID"])
        .reset_index(drop=True)
    )

    for mkt in event_df["Market"].unique():
        mask = event_df["Market"] == mkt
        event_df.loc[mask, "Back_Cumulative_PnL"] = (
            event_df.loc[mask, "Back_PnL"].cumsum().round(2).to_numpy()
        )

    season_rows = []
    for test_year in sorted(pred_df["Test_Year"].unique()):
        for market_name, market_config in BETTING_MARKETS.items():
            target_col = market_config["target_col"]
            mdf = pred_df[
                (pred_df["Test_Year"] == test_year) &
                (pred_df["Market"]    == market_name)
            ]
            if len(mdf) == 0 or target_col not in mdf.columns:
                continue
            y_true = mdf["Actual"].values
            y_prob = mdf["Normalised_Probability"].values
            disc = compute_discrimination(y_true, y_prob)
            bs   = back_summary(mdf, target_col)
            season_rows.append({
                "Tour":       tour_key,
                "Test_Year":  test_year,
                "Market":     market_name,
                "N_Events":   int(mdf["EventID"].nunique()),
                "N_Players":  len(mdf),
                "Prevalence": round(float(y_true.mean()), 4),
                **disc, **bs,
            })

    summary_rows = []
    calib_sheets = {}
    for market_name, market_config in BETTING_MARKETS.items():
        target_col = market_config["target_col"]
        mdf = pred_df[pred_df["Market"] == market_name]
        if len(mdf) == 0 or target_col not in mdf.columns:
            continue
        y_true = mdf["Actual"].values
        y_prob = mdf["Normalised_Probability"].values
        disc = compute_discrimination(y_true, y_prob)
        bs   = back_summary(mdf, target_col)
        summary_rows.append({
            "Tour":          tour_key,
            "Market":        market_name,
            "N_Test_Years":  int(pred_df[pred_df["Market"] == market_name]["Test_Year"].nunique()),
            "N_Events":      int(mdf["EventID"].nunique()),
            "N_Players":     len(mdf),
            "Prevalence":    round(float(y_true.mean()), 4),
            **disc, **bs,
        })
        calib_sheets[market_name] = calibration_bins(y_true, y_prob)

    return {
        "summary":         pd.DataFrame(summary_rows),
        "season_summary":  pd.DataFrame(season_rows),
        "event_results":   event_df,
        "all_predictions": pred_df,
        "calibration":     calib_sheets,
    }


# ===== GRID SEARCH =====

def run_grid_search(pred_df: pd.DataFrame) -> pd.DataFrame:
    """
    Sweep edge/odds/rating filters across the combined walk-forward predictions.

    Edge is defined as Lay_odds / Normalised_Model_Odds. P&L uses the same
    formula-based calculation as apply_back_strategy (Lay odds, no dead heat).
    Results are sorted by ROI%.
    """
    df = pred_df.copy()

    # Resolve which Lay odds column applies to each row
    df["_Lay_Odds"] = np.nan
    for market_name, lay_col in LAY_ODDS_COLS.items():
        if lay_col in df.columns:
            mask = df["Market"] == market_name
            df.loc[mask, "_Lay_Odds"] = df.loc[mask, lay_col]

    df["_Edge"] = df["_Lay_Odds"] / df["Normalised_Model_Odds"].clip(1e-8)

    # Formula-based unit P&L with dead-heat RF for place markets.
    # RF is already stored on each row from apply_back_strategy; fall back to
    # 1.0 if the column is absent (e.g. Winner market rows).
    actual = df["Actual"].to_numpy(dtype=float)
    lay    = df["_Lay_Odds"].to_numpy(dtype=float)
    rf     = df["DeadHeat_RF"].to_numpy(dtype=float) if "DeadHeat_RF" in df.columns else np.ones(len(df))
    df["_Unit_PnL"] = np.where(
        actual == 1, BACK_STAKE * (rf * lay - 1), -BACK_STAKE
    )

    market_dfs = {m: df[df["Market"] == m] for m in BETTING_MARKETS}

    combos = [
        (mkt, edge, mn_o, mx_o, mn_r)
        for mkt, edge, mn_o, mx_o, mn_r
        in product(list(BETTING_MARKETS.keys()),
                   GRID_EDGE_THRESHOLDS, GRID_MIN_ODDS, GRID_MAX_ODDS, GRID_MIN_RATING)
        if mn_o < mx_o
    ]
    total = len(combos)
    print(f"\n  Grid search: {total:,} combinations across "
          f"{len(pred_df):,} predictions...")

    results = []
    for i, (mkt, edge, mn_o, mx_o, mn_r) in enumerate(combos):
        if mkt not in market_dfs:
            continue
        sub  = market_dfs[mkt]
        mask = (sub["_Edge"] >= edge) & (sub["_Lay_Odds"] >= mn_o) & (sub["_Lay_Odds"] <= mx_o)
        if mn_r is not None:
            mask = mask & (sub["rating"] >= mn_r)
        filtered = sub[mask]
        if len(filtered) == 0:
            continue

        pnl_vals     = filtered["_Unit_PnL"].values
        actuals      = filtered["Actual"].values
        odds         = filtered["_Lay_Odds"].values
        n_bets       = len(filtered)
        total_staked = n_bets * BACK_STAKE
        total_pnl    = pnl_vals.sum()
        roi          = total_pnl / total_staked if total_staked > 0 else 0

        event_pnl = pd.Series(pnl_vals, index=filtered.index).groupby(
            filtered["EventID"]).sum()
        n_events  = len(event_pnl)
        epnl_std  = event_pnl.std()
        sharpe    = (event_pnl.mean() / epnl_std * np.sqrt(n_events)) if epnl_std > 0 else 0
        cum       = event_pnl.cumsum().values
        max_dd    = (cum - np.maximum.accumulate(cum)).min()

        results.append({
            "Market":         mkt,
            "Edge_Threshold": edge,
            "Min_Odds":       mn_o,
            "Max_Odds":       mx_o if mx_o < 9999 else "None",
            "Min_Rating":     mn_r if mn_r is not None else "None",
            "N_Bets":         n_bets,
            "N_Won":          int(actuals.sum()),
            "Strike_Rate_%":  round(actuals.mean() * 100, 2),
            "Total_Staked":   round(total_staked, 2),
            "Total_PnL":      round(total_pnl, 2),
            "ROI_%":          round(roi * 100, 2),
            "Avg_Lay_Odds":   round(odds.mean(), 2),
            "Sharpe":         round(sharpe, 3),
            "Max_Drawdown":   round(max_dd, 2),
        })

        if (i + 1) % 1000 == 0:
            print(f"    {i+1:,}/{total:,} done, {len(results):,} valid so far...")

    if not results:
        return pd.DataFrame()

    grid_df = pd.DataFrame(results).sort_values("ROI_%", ascending=False).reset_index(drop=True)
    n_profitable = (grid_df["Total_PnL"] > 0).sum()
    print(f"  Grid complete: {len(grid_df):,} valid strategies, "
          f"{n_profitable:,} profitable ({100*n_profitable/len(grid_df):.1f}%)")
    print(f"  Best ROI: {grid_df['ROI_%'].max():.1f}%  |  "
          f"Best P&L: £{grid_df['Total_PnL'].max():,.0f}")
    return grid_df


def _write_grid_sheet(wb, grid_df: pd.DataFrame, sheet_name: str = "Strategy_Grid"):
    """Write the grid search results to a styled Excel sheet."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    thin_side  = Side(style="thin", color="CCCCCC")
    std_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center     = Alignment(horizontal="center", vertical="center")
    HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HDR_FILL   = PatternFill("solid", start_color="1F4E79")
    BODY_FONT  = Font(name="Arial", size=9)
    POS_FILL   = PatternFill("solid", start_color="C6EFCE")
    NEG_FILL   = PatternFill("solid", start_color="FFC7CE")
    MID_FILL   = PatternFill("solid", start_color="FFEB9C")

    cols = list(grid_df.columns)
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = center
        cell.border    = std_border

    pnl_col_idx = cols.index("Total_PnL") + 1
    roi_col_idx = cols.index("ROI_%") + 1

    for row_idx, row in grid_df.iterrows():
        excel_row = row_idx + 2
        for col_idx, val in enumerate(row.values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.border    = std_border
            cell.alignment = center
        pnl = row["Total_PnL"]
        ws.cell(row=excel_row, column=pnl_col_idx).fill = (
            POS_FILL if pnl > 0 else NEG_FILL if pnl < -200 else MID_FILL
        )
        ws.cell(row=excel_row, column=roi_col_idx).fill = (
            POS_FILL if row["ROI_%"] > 0 else NEG_FILL
        )

    col_widths = {
        "Market": 10, "Edge_Threshold": 15, "Min_Odds": 11, "Max_Odds": 11,
        "Min_Rating": 12, "N_Bets": 9, "N_Won": 8, "Strike_Rate_%": 14,
        "Total_Staked": 14, "Total_PnL": 13, "ROI_%": 9,
        "Avg_Lay_Odds": 13, "Sharpe": 10, "Max_Drawdown": 15,
    }
    for col_idx, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 12)

    ws.freeze_panes = "A2"


# ===== EXPORT =====

def export_results(results: dict, tour_key: str, grid_df: pd.DataFrame = None) -> Path:
    WF_RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = WF_RESULTS_DIR / f"{tour_key}_WalkForward_Backtest.xlsx"

    id_cols   = [c for c in ["Test_Year", "Date", "EventID", "Market",
                              "surname", "firstname", "posn", "rating"]
                 if c in results["all_predictions"].columns]
    lay_cols  = [c for c in LAY_ODDS_COLS.values()
                 if c in results["all_predictions"].columns]
    pred_cols = ["Model_Score", "Probability", "Normalised_Probability",
                 "Normalised_Model_Odds", "Actual"]
    bet_cols  = ["Back_Bet", "DeadHeat_RF", "Back_PnL"]
    all_cols  = id_cols + lay_cols + pred_cols + bet_cols
    export_df = results["all_predictions"][
        [c for c in all_cols if c in results["all_predictions"].columns]
    ]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        results["summary"].to_excel(        writer, sheet_name="Summary",         index=False)
        results["season_summary"].to_excel( writer, sheet_name="Season_Summary",  index=False)
        results["event_results"].to_excel(  writer, sheet_name="Event_Results",   index=False)
        export_df.to_excel(                 writer, sheet_name="All_Predictions", index=False)
        for mkt, calib_df in results["calibration"].items():
            calib_df.to_excel(writer, sheet_name=f"Calib_{mkt}"[:31], index=False)

    if grid_df is not None and len(grid_df) > 0:
        wb = load_workbook(out_path)
        _write_grid_sheet(wb, grid_df)
        wb.save(out_path)

    print(f"\n  Saved: {out_path}")
    return out_path


# ===== TOUR RUNNER =====

def run_tour(tour_key: str, tour_info: dict, n_trials: int,
             min_test_year: int, force_retrain: bool):
    print(f"\n{'='*60}")
    print(f"WALK-FORWARD: {tour_info['name']}")
    print(f"{'='*60}")

    hist_path = tour_info["historical_file"]
    if not hist_path.exists():
        print(f"  Processed file not found: {hist_path}")
        return None

    df = pd.read_excel(hist_path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.dropna(subset=["Date"])

    # Warn if any Lay odds columns are missing from the data
    missing_lay = [col for col in LAY_ODDS_COLS.values() if col not in df.columns]
    if missing_lay:
        print(f"  WARNING: Lay odds columns not found in data: {missing_lay}")
        print(f"  Affected markets will be skipped during backtesting.")

    print(f"  Loaded {len(df):,} rows  |  years: "
          f"{int(df['Date'].dt.year.min())}–{int(df['Date'].dt.year.max())}")

    windows = get_windows(df, min_test_year=min_test_year)
    if not windows:
        print("  No valid walk-forward windows found."); return None

    print(f"\n  Walk-forward plan ({len(windows)} windows):")
    for ts, te, ty in windows:
        cached = (WF_MODELS_DIR / f"{tour_key}_{ty}" / "market_bundle.pkl").exists()
        status = " [cached]" if cached and not force_retrain else ""
        print(f"    Train {ts}–{te}  →  Test {ty}{status}")

    all_preds_list  = []
    event_summ_list = []
    prev_window_dir = None

    for train_start, train_end, test_year in windows:
        print(f"\n  --- Train {train_start}–{train_end} → Test {test_year} ---")

        bundle_path = WF_MODELS_DIR / f"{tour_key}_{test_year}" / "market_bundle.pkl"

        if bundle_path.exists() and not force_retrain:
            print(f"  Loading cached bundle for {test_year}...")
            package = joblib.load(bundle_path)
        else:
            train_df = df[
                (df["Date"].dt.year >= train_start) &
                (df["Date"].dt.year <= train_end)
            ].copy()
            print(f"  Training rows: {len(train_df):,}")

            package = train_window(
                tour_key, train_df, test_year, n_trials, prev_window_dir
            )
            if not package:
                print(f"  No markets trained — skipping test year {test_year}")
                prev_window_dir = WF_MODELS_DIR / f"{tour_key}_{test_year}"
                continue

            bundle_path.parent.mkdir(parents=True, exist_ok=True)
            joblib.dump(package, bundle_path)
            print(f"  Bundle saved: {bundle_path.parent.name}/market_bundle.pkl")

        prev_window_dir = WF_MODELS_DIR / f"{tour_key}_{test_year}"

        test_df = df[df["Date"].dt.year == test_year].copy()
        if len(test_df) == 0:
            print(f"  No test data for {test_year}"); continue

        window_preds, window_events = backtest_window(
            tour_key, test_df, test_year, package
        )
        all_preds_list.extend(window_preds)
        event_summ_list.extend(window_events)

    results = aggregate_results(all_preds_list, event_summ_list, tour_key)
    if results is None:
        print("  No results to aggregate."); return None

    print(f"\n  === AGGREGATE WALK-FORWARD RESULTS ({len(windows)} windows) ===")
    print(f"  {'Market':<8}  {'AUC':>6}  {'AP':>6}  {'TSS':>6}  {'Back_PnL':>10}  {'Back_ROI':>9}")
    for _, row in results["summary"].iterrows():
        print(
            f"  {row['Market']:<8}  {row['AUC']:>6.4f}  {row['Avg_Precision']:>6.4f}  "
            f"{row['TSS']:>6.4f}  £{row['Back_PnL']:>9.2f}  {row['Back_ROI']:>8.1%}"
        )

    grid_df = run_grid_search(results["all_predictions"])

    return results, grid_df


# ===== MAIN =====

def main():
    args = parse_args()

    print("=== GOLF MODEL WALK-FORWARD BACKTESTING ===")
    print(f"Training window: {TRAINING_YEARS} years  |  Optuna trials: {args.trials}")
    print(f"Back stake: £{BACK_STAKE}  |  Lay odds P&L  |  Dead-heat RF applied to place markets")
    if args.force_retrain:
        print("  ** Force retrain: ignoring all cached models **")

    tours_to_run = {
        k: v for k, v in TOUR_CONFIG.items()
        if args.tour is None or k == args.tour
    }

    for tour_key, tour_info in tours_to_run.items():
        outcome = run_tour(
            tour_key, tour_info,
            n_trials=args.trials,
            min_test_year=args.min_year,
            force_retrain=args.force_retrain,
        )
        if outcome:
            results, grid_df = outcome
            export_results(results, tour_key, grid_df)


if __name__ == "__main__":
    main()