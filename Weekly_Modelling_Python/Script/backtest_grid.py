"""
PGA Back Betting Strategy Grid Search
======================================
Sweeps edge thresholds, odds filters, rating filters, and staking methods
across all four markets. Writes results to a new sheet in the workbook.

Usage:
    python backtest_grid.py
    python backtest_grid.py path/to/PGA_WalkForward_Backtest.xlsx
"""

import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from itertools import product
import warnings
warnings.filterwarnings("ignore")

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_FILE   = r"D:\Golf\Repo\Weekly_Modelling_Python\Output\WalkForward\Results\PGA_WalkForward_Backtest.xlsx"
OUTPUT_SHEET = "Strategy_Grid"
FIXED_STAKE   = 10
BANKROLL     = 2000
KELLY_FRAC   = 0.25

# Grid dimensions
EDGE_THRESHOLDS = [1.0, 1.05, 1.1, 1.25, 1.5, 2.0, 3.0, 5.0]
MIN_ODDS        = [1.0, 3.0, 5.0, 10.0]
MAX_ODDS        = [9999, 10, 25, 50, 100, 200, 500]   # 9999 = no cap
MIN_RATING      = [None, 50, 55, 60, 65, 70]           # None = no filter
MARKETS         = ["Winner", "Top5", "Top10", "Top20"]
STAKING_METHODS = ["FIXED"]

# ── Load & prepare data ───────────────────────────────────────────────────────
print(f"Loading {INPUT_FILE}...")
df = pd.read_excel(INPUT_FILE, sheet_name="All_Predictions")

conditions = [
    df["Market"] == "Winner",
    df["Market"] == "Top5",
    df["Market"] == "Top10",
]
choices = [df["Win_odds"], df["Top5_odds"], df["Top10_odds"]]
df["Market_Odds"] = np.select(conditions, choices, default=df["Top20_odds"])
df["Edge"] = df["Market_Odds"] / df["Normalised_Model_Odds"]

market_dfs = {m: df[df["Market"] == m].copy() for m in MARKETS}
print(f"Data loaded: {len(df):,} rows across {len(MARKETS)} markets")

# ── Core strategy evaluator (vectorised) ─────────────────────────────────────
def run_strategy(sub, staking):
    if len(sub) == 0:
        return None

    if staking == "FIXED":
        stakes = np.full(len(sub), float(FIXED_STAKE))

    elif staking == "Kelly":
        p = sub["Normalised_Probability"].values
        b = sub["Market_Odds"].values - 1
        k = np.where(b > 0, (b * p - (1 - p)) / b, 0)
        stakes = np.clip(KELLY_FRAC * k * BANKROLL, 0, BANKROLL * 0.05)
        stakes = np.round(stakes, 2)

    else:  # Edge_Proportional
        stakes = np.round(np.clip(FIXED_STAKE * (sub["Edge"].values - 1), 0, FIXED_STAKE * 10), 2)

    mask = stakes > 0
    if mask.sum() == 0:
        return None

    stakes  = stakes[mask]
    actuals = sub["Actual"].values[mask]
    odds    = sub["Market_Odds"].values[mask]

    pnl = np.where(actuals == 1, stakes * (odds - 1), -stakes)

    n_bets       = len(stakes)
    total_staked = stakes.sum()
    total_pnl    = pnl.sum()
    roi          = total_pnl / total_staked if total_staked > 0 else 0

    # Sharpe and drawdown aggregated to event level
    event_pnl   = pd.Series(pnl, index=sub[mask].index).groupby(
                      sub[mask]["EventID"]).sum()
    n_events    = len(event_pnl)
    epnl_std    = event_pnl.std()
    sharpe      = (event_pnl.mean() / epnl_std * np.sqrt(n_events)) if epnl_std > 0 else 0
    cum         = event_pnl.cumsum().values
    max_dd      = (cum - np.maximum.accumulate(cum)).min()

    return {
        "N_Bets":        n_bets,
        "N_Won":         int(actuals.sum()),
        "Strike_Rate_%": round(actuals.mean() * 100, 2),
        "Total_Staked":  round(total_staked, 2),
        "Total_PnL":     round(total_pnl, 2),
        "ROI_%":         round(roi * 100, 2),
        "Avg_Odds":      round(odds.mean(), 2),
        "Sharpe":        round(sharpe, 3),
        "Max_Drawdown":  round(max_dd, 2),
    }

# ── Grid search ───────────────────────────────────────────────────────────────
print("Running grid search...")
results = []

combos = [
    (mkt, edge, mn_o, mx_o, mn_r, stk)
    for mkt, edge, mn_o, mx_o, mn_r, stk
    in product(MARKETS, EDGE_THRESHOLDS, MIN_ODDS, MAX_ODDS, MIN_RATING, STAKING_METHODS)
    if mn_o < mx_o
]
total = len(combos)
print(f"  {total:,} combinations to evaluate")

for i, (mkt, edge, mn_o, mx_o, mn_r, stk) in enumerate(combos):
    sub = market_dfs[mkt]
    mask = (sub["Edge"] >= edge) & (sub["Market_Odds"] >= mn_o) & (sub["Market_Odds"] <= mx_o)
    if mn_r is not None:
        mask &= sub["rating"] >= mn_r
    metrics = run_strategy(sub[mask], stk)

    if metrics:
        results.append({
            "Market":         mkt,
            "Edge_Threshold": edge,
            "Min_Odds":       mn_o,
            "Max_Odds":       mx_o if mx_o < 9999 else "None",
            "Min_Rating":     mn_r if mn_r is not None else "None",
            **metrics,
        })

    if (i + 1) % 1000 == 0:
        print(f"  {i+1:,}/{total:,} done, {len(results):,} valid strategies...")

results_df = pd.DataFrame(results).sort_values("ROI_%", ascending=False).reset_index(drop=True)
n_profitable = (results_df["Total_PnL"] > 0).sum()

print(f"\nGrid complete: {len(results_df):,} valid strategies")
print(f"Profitable (PnL > 0): {n_profitable:,} ({100*n_profitable/len(results_df):.1f}%)")
print(f"Best ROI:  {results_df['ROI_%'].max():.1f}%")
print(f"Best PnL:  £{results_df['Total_PnL'].max():,.0f}")
print(f"\nTop 5 by ROI%:")
print(results_df[["Market","Staking","Edge_Threshold","Min_Odds","Max_Odds",
                   "Min_Rating","N_Bets","ROI_%","Total_PnL","Sharpe"]].head().to_string(index=False))

# ── Write to workbook ─────────────────────────────────────────────────────────
print(f"\nWriting to '{OUTPUT_SHEET}' sheet in {INPUT_FILE}...")
wb = load_workbook(INPUT_FILE)
if OUTPUT_SHEET in wb.sheetnames:
    del wb[OUTPUT_SHEET]
ws = wb.create_sheet(OUTPUT_SHEET)

thin_side  = Side(style="thin", color="CCCCCC")
std_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
center     = Alignment(horizontal="center", vertical="center")

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
BODY_FONT = Font(name="Arial", size=9)
POS_FILL  = PatternFill("solid", start_color="C6EFCE")
NEG_FILL  = PatternFill("solid", start_color="FFC7CE")
MID_FILL  = PatternFill("solid", start_color="FFEB9C")
META_FILL = PatternFill("solid", start_color="D9E1F2")

# Column headers (row 4)
cols = list(results_df.columns)
for col_idx, col_name in enumerate(cols, 1):
    cell = ws.cell(row=4, column=col_idx, value=col_name)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = center
    cell.border = std_border
ws.row_dimensions[1].height = 18

# Data rows
pnl_col_idx = cols.index("Total_PnL") + 1
roi_col_idx = cols.index("ROI_%") + 1

for row_idx, row in results_df.iterrows():
    excel_row = row_idx + 5
    for col_idx, val in enumerate(row.values, 1):
        cell = ws.cell(row=excel_row, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.border = std_border
        cell.alignment = center

    pnl = row["Total_PnL"]
    ws.cell(row=excel_row, column=pnl_col_idx).fill = (
        POS_FILL if pnl > 0 else NEG_FILL if pnl < -200 else MID_FILL
    )
    ws.cell(row=excel_row, column=roi_col_idx).fill = (
        POS_FILL if row["ROI_%"] > 0 else NEG_FILL
    )

# Column widths
widths = {
    "Market": 10, "Staking": 20, "Edge_Threshold": 15, "Min_Odds": 11,
    "Max_Odds": 11, "Min_Rating": 12, "N_Bets": 9, "N_Won": 8,
    "Strike_Rate_%": 14, "Total_Staked": 14, "Total_PnL": 13,
    "ROI_%": 9, "Avg_Odds": 11, "Sharpe": 10, "Max_Drawdown": 15,
}
for col_idx, col_name in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 12)

ws.freeze_panes = "A5"

wb.save(INPUT_FILE)
print(f"Done. Open {INPUT_FILE} and navigate to the '{OUTPUT_SHEET}' sheet.")