"""
Walk-Forward Backtesting Script

Trains on a rolling 2-year window, predicts on the next year, slides forward
through all available data. No pre-trained models required — this script trains
and tests every window from scratch (with warm-starting between adjacent windows).

Example: data available 2018–2026
  Window 1:  Train 2018–2019  →  Test 2020
  Window 2:  Train 2019–2020  →  Test 2021
  Window 3:  Train 2020–2021  →  Test 2022
  Window 4:  Train 2021–2022  →  Test 2023
  Window 5:  Train 2022–2023  →  Test 2024
  Window 6:  Train 2023–2024  →  Test 2025
  Window 7:  Train 2024–2025  →  Test 2026

Odds
----
All P&L calculations use Betfair Lay odds (Lay_odds, Lay_top5, Lay_top10,
Lay_top20), which are pre-event snapshots and closer to expected back prices
than bookie odds. No pre-computed profit columns are used; P&L is formula-based
for all markets with dead-heat adjustment applied to place markets.

Dead heat (place markets only)
-------------------------------
A dead heat occurs when players tie at exactly the cut position (5, 10, or 20),
meaning fewer places are available than players tied there. Only players AT the
cut position are affected; players finishing clearly inside the cut get full odds.

  places_filled_above     = count of players with posn < cut
  places_available_in_tie = cut - places_filled_above
  players_tied            = count of players with posn == cut position in data
  RF                      = places_available_in_tie / players_tied  (clamped 0–1)

  Win P&L  = stake × (RF × lay_odds - 1)
  Loss P&L = -stake

Caching
-------
Trained model bundles are saved to Output/WalkForward/Models/{tour}_{year}/
so the script can be interrupted and resumed without retraining completed windows.
Use --force-retrain to ignore the cache and retrain everything.

Optuna trials
-------------
Default is 30 trials per model per market (vs 75 in production) to keep runtimes
manageable across many windows. Use --trials 75 for full production quality.
Warm-start: best params from window N are passed as starting suggestions to window N+1.

Back betting
------------
Bet when Normalised_Model_Odds < Lay_odds (fixed £10 stake).
Winner:          P&L = (lay_odds - 1) × 10 on win, -10 on loss.
Top5/Top10/Top20: same formula with dead-heat RF applied to winning rows.

Grid search
-----------
After all windows are processed a parameter grid is swept over the combined
All_Predictions data. Dimensions:
  Edge threshold  — ratio of lay odds to model odds required before betting
  Min/Max odds    — odds range filter
  Min rating      — player rating floor filter
Results are written to the Strategy_Grid sheet of the output workbook.

Output
------
Output/WalkForward/Results/{tour}_WalkForward_Backtest.xlsx
  Summary         — overall metrics + back P&L across all windows, per market
  Season_Summary  — same broken down by test year
  Event_Results   — per-event P&L with cumulative P&L column
  All_Predictions — every player prediction + outcome + bet P&L
  Calib_*         — calibration bins (predicted vs actual) per market
  Strategy_Grid   — grid search results sorted by ROI

Run:
  python walk_forward_backtest.py                         # both tours, all windows, 30 trials
  python walk_forward_backtest.py --tour PGA
  python walk_forward_backtest.py --trials 75             # full Optuna
  python walk_forward_backtest.py --start-year 2022       # discard data before 2022 (faster)
  python walk_forward_backtest.py --min-year 2024         # skip windows with test year < 2024
  python walk_forward_backtest.py --force-retrain         # ignore cached models

  --start-year trims the data loaded into memory; combined with --min-year you can
  run a single test window quickly, e.g.:
    --start-year 2022 --min-year 2024   loads 2022+, only evaluates 2024 test window
"""

import argparse
import shutil
import sys
import warnings
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import (
    average_precision_score,
    brier_score_loss,
    log_loss,
    roc_auc_score,
)

import seasonal_model_training as smt

from config import (
    BASE_DIR,
    BETTING_MARKETS,
    PREDICTIONS_DIR,
    TOUR_CONFIG,
    TRAINING_YEARS,
)
from seasonal_model_training import get_market_vars, tss_optimal

warnings.filterwarnings("ignore")

# ===== PATHS =====
WF_DIR         = BASE_DIR / "Output" / "WalkForward"
WF_MODELS_DIR  = WF_DIR / "Models"
WF_RESULTS_DIR = WF_DIR / "Results"

# ===== CONSTANTS =====
BACK_STAKE        = 10.0   # £ per back bet
DEFAULT_WF_TRIALS = 30
MIN_POSITIVES     = 10     # skip a market window if training set has fewer positives

# ===== LAY STAKING =====
# Fixed liability per player per market (£)
LAY_FIXED_LIABILITY = {
    "Winner": 1000.0,
    "Top5":    200.0,
    "Top10":   100.0,
    "Top20":    50.0,
}

# Fixed stake per player per market (£)
LAY_FIXED_STAKE = {
    "Winner":  1.0,
    "Top5":    5.0,
    "Top10":  10.0,
    "Top20":  20.0,
}

# ===== LAY ODDS COLUMN MAPPING =====
# Maps each market name to its Betfair Lay odds column.
# These are pre-event snapshots used for both the edge condition and P&L.
LAY_ODDS_COLS = {
    "Winner": "Lay_odds",
    "Top5":   "Lay_top5",
    "Top10":  "Lay_top10",
    "Top20":  "Lay_top20",
}

# Cut position for each place market — used for dead-heat calculation.
# Winner has no cut position (no dead-heat logic applied).
PLACE_MARKET_CUTS = {
    "Top5":  5,
    "Top10": 10,
    "Top20": 20,
}

# ===== GRID SEARCH PARAMETERS =====
# Back grid: floor thresholds — bet if rating >= threshold
# Lay grid:  ceiling thresholds — bet if rating < threshold
# "All" row (no filter) is always included as baseline in both grids.
GRID_BACK_RATING_FLOORS   = [55, 60, 65, 70]
GRID_LAY_RATING_CEILINGS  = [55, 60, 65, 70]


# ===== CLI =====

def parse_args():
    parser = argparse.ArgumentParser(
        description="Walk-forward backtest for golf prediction models"
    )
    parser.add_argument("--tour", default=None,
                        help="Tour to backtest: PGA or Euro (default: both)")
    parser.add_argument("--trials", type=int, default=DEFAULT_WF_TRIALS,
                        help=f"Optuna trials per model per market "
                             f"(default: {DEFAULT_WF_TRIALS})")
    parser.add_argument("--start-year", type=int, default=None, dest="start_year",
                        help="Discard all data before this year (reduces memory and load time)")
    parser.add_argument("--min-year", type=int, default=None, dest="min_year",
                        help="Earliest test year to evaluate (default: all available)")
    parser.add_argument("--force-retrain", action="store_true", dest="force_retrain",
                        help="Ignore cached models and retrain all windows")
    return parser.parse_args()


# ===== PREDICTION =====

def predict_event(event_df: pd.DataFrame, market_name: str,
                  market_pkg: dict) -> pd.DataFrame | None:
    """
    Apply a trained market package to a single event's player rows.
    Mirrors the weekly prediction script exactly so backtest reflects
    real deployment behaviour.
    """
    model_vars = market_pkg["model_vars"]
    odds_col   = market_pkg["odds_col"]

    available = [v for v in model_vars if v in event_df.columns]
    df = event_df[available + [odds_col]].copy()
    for col in available + [odds_col]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna()
    if len(df) == 0:
        return None

    X    = df[available].values.astype(float)
    odds = df[odds_col].values.astype(float)

    model_preds = np.column_stack([
        m.predict_proba(X)[:, 1] for m in market_pkg["models"].values()
    ])
    raw_score = model_preds.mean(axis=1)

    imp_prob      = (1.0 / odds.clip(1e-8)).reshape(-1, 1)
    meta_X_scaled = market_pkg["meta_scaler"].transform(np.hstack([model_preds, imp_prob]))
    proba         = market_pkg["meta_model"].predict_proba(meta_X_scaled)[:, 1]

    market_size = market_pkg["market_size"]
    prob_sum    = proba.sum()
    norm_prob   = (proba / prob_sum) * market_size if prob_sum > 0 else proba

    result = event_df.loc[df.index].copy()
    result["Model_Score"]            = raw_score.round(5)
    result["Probability"]            = proba.round(6)
    result["Normalised_Probability"] = norm_prob.round(6)
    result["Normalised_Model_Odds"]  = (1.0 / norm_prob.clip(1e-8)).round(2)
    return result


# ===== DEAD HEAT =====

def compute_reduction_factors(event_df: pd.DataFrame, cut: int) -> pd.Series:
    """
    Compute a per-row dead-heat reduction factor (RF) for a place market.

    Only rows where posn == the tied cut position receive RF < 1.
    Rows finishing clearly inside the cut (posn < cut) always get RF = 1.0
    because they are unambiguous winners regardless of any tie at the boundary.

    Parameters
    ----------
    event_df : DataFrame for a single event (all players).
    cut      : Cut position for the market (5, 10, or 20).

    Returns
    -------
    pd.Series of RF values aligned to event_df.index.
    """
    posn = pd.to_numeric(event_df["posn"], errors="coerce")
    rf   = pd.Series(1.0, index=event_df.index)

    qualifying = posn[posn <= cut]
    if len(qualifying) == 0:
        return rf

    # The actual boundary position (e.g. 9 when 11 players tie for 9th)
    tied_posn           = int(qualifying.max())
    places_filled_above = int((posn < tied_posn).sum())
    places_available    = cut - places_filled_above
    tied_mask           = posn == tied_posn
    players_tied        = int(tied_mask.sum())

    if players_tied > 1 and places_available < players_tied:
        # Genuine dead heat — clamp RF to [0, 1] as a safety guard
        raw_rf = places_available / players_tied
        dh_rf  = float(np.clip(raw_rf, 0.0, 1.0))
        rf[tied_mask] = dh_rf

    return rf


# ===== BACK BETTING STRATEGY =====

def apply_back_strategy(df: pd.DataFrame, lay_odds_col: str,
                        target_col: str,
                        market_name: str = "Winner") -> pd.DataFrame:
    """
    Back when Normalised_Model_Odds < Lay_odds (model thinks player underpriced).
    Fixed stake BACK_STAKE per bet.

    Winner market — straight formula, no dead-heat adjustment:
      Win:  P&L = (lay_odds - 1) * BACK_STAKE
      Loss: P&L = -BACK_STAKE

    Place markets (Top5 / Top10 / Top20) — dead-heat RF applied to winners
    who are tied exactly at the cut position:
      Win (clear):       P&L = (lay_odds - 1) * BACK_STAKE          [RF = 1]
      Win (dead heat):   P&L = BACK_STAKE * (RF * lay_odds - 1)     [RF < 1]
      Loss:              P&L = -BACK_STAKE

    The RF is computed per-event from the posn column. If posn is missing
    the function falls back to RF = 1.0 (no adjustment) for that event.
    """
    df     = df.copy()
    lay    = df[lay_odds_col].to_numpy(dtype=float)
    actual = df[target_col].to_numpy(dtype=float)
    model  = df["Normalised_Model_Odds"].to_numpy(dtype=float)
    is_back = model < lay

    # Build RF array — default 1.0 (no dead heat)
    rf = np.ones(len(df), dtype=float)

    cut = PLACE_MARKET_CUTS.get(market_name)
    if cut is not None and "posn" in df.columns:
        # Compute RF per event so ties are evaluated within each field
        for event_id, idx in df.groupby("eventID").groups.items():
            event_slice = df.loc[idx]
            event_rf    = compute_reduction_factors(event_slice, cut)
            rf[df.index.get_indexer(idx)] = event_rf.values

    # P&L: stake × (RF × odds - 1) on win, -stake on loss
    # When RF == 1 this reduces to the standard (odds - 1) × stake formula
    back_pnl = np.where(
        is_back & (actual == 1),  BACK_STAKE * (rf * lay - 1),
        np.where(is_back, -BACK_STAKE, 0.0)
    )

    df["Back_Bet"]          = is_back
    df["Back_PnL"]          = back_pnl
    df["DeadHeat_RF"]       = rf          # retained for inspection / audit
    return df


# ===== STRATEGY SUMMARY =====

def back_summary(df: pd.DataFrame, target_col: str) -> dict:
    bets = df[df["Back_Bet"]]
    if len(bets) == 0:
        return {"Back_NBets": 0, "Back_NWon": 0, "Back_HitRate": np.nan,
                "Back_PnL": 0.0, "Back_ROI": np.nan}
    n_bets = len(bets)
    n_won  = int(bets[target_col].sum())
    pnl    = float(bets["Back_PnL"].sum())
    return {
        "Back_NBets":   n_bets,
        "Back_NWon":    n_won,
        "Back_HitRate": round(n_won / n_bets, 4),
        "Back_PnL":     round(pnl, 2),
        "Back_ROI":     round(pnl / (n_bets * BACK_STAKE), 4),
    }


# ===== LAY BETTING STRATEGY =====

def _lay_return(lay: np.ndarray, actual: np.ndarray, rf: np.ndarray,
                exp_winners: np.ndarray, act_winners: np.ndarray,
                liability: np.ndarray) -> np.ndarray:
    """
    Core dead-heat aware lay return formula (vectorised).

    For winning outcomes (actual == 1, i.e. bad for the layer):
        Return = ((1 - (exp_winners / act_winners) * lay) / (lay - 1)) * liability

    For losing outcomes (actual == 0, i.e. good for the layer):
        Return = liability / (lay - 1)    [= stake]

    When there is no dead heat exp_winners == act_winners == 1, so the formula
    reduces to the standard -liability (win) / +stake (loss).

    exp_winners : places available in the tie  (= places_available from RF logic)
    act_winners : players tied at the cut      (= players_tied from RF logic)
    Both are 1.0 for no-dead-heat rows and for the Winner market.
    """
    stake      = liability / np.clip(lay - 1, 1e-8, None)
    win_return = ((1 - (exp_winners / np.clip(act_winners, 1e-8, None)) * lay)
                  / np.clip(lay - 1, 1e-8, None)) * liability
    los_return = stake
    return np.where(actual == 1, win_return, los_return)


def _build_dh_arrays(df: pd.DataFrame, market_name: str):
    """
    Return (exp_winners, act_winners) arrays aligned to df.index.

    For the Winner market or when posn is absent both arrays are all 1.0
    (no dead heat possible).  For place markets we re-use the same
    compute_reduction_factors logic to derive the raw counts.
    """
    n           = len(df)
    exp_arr     = np.ones(n, dtype=float)
    act_arr     = np.ones(n, dtype=float)
    cut         = PLACE_MARKET_CUTS.get(market_name)

    if cut is None or "posn" not in df.columns:
        return exp_arr, act_arr

    posn = pd.to_numeric(df["posn"], errors="coerce")

    for _, idx in df.groupby("eventID").groups.items():
        ep   = posn.loc[idx]
        qual = ep[ep <= cut]
        if len(qual) == 0:
            continue
        tied_posn            = qual.max()
        places_filled_above  = int((ep < tied_posn).sum())
        places_available     = cut - places_filled_above
        players_tied         = int((ep == tied_posn).sum())

        if players_tied > places_available:
            # Dead heat — store counts for affected rows
            tied_iloc = [df.index.get_loc(i) for i in idx if posn.loc[i] == tied_posn]
            exp_arr[tied_iloc] = float(places_available)
            act_arr[tied_iloc] = float(players_tied)
        # else: no dead heat, arrays stay 1.0

    return exp_arr, act_arr


def apply_lay_strategy(df: pd.DataFrame, lay_odds_col: str,
                       target_col: str,
                       market_name: str = "Winner") -> pd.DataFrame:
    """
    Lay when Normalised_Model_Odds > Lay_odds (model thinks player overpriced).

    Two staking methods computed in parallel:

    Fixed liability (LAY_FIXED_LIABILITY per market):
        Stake  = liability / (lay_odds - 1)
        Return = dead-heat formula applied to fixed liability

    Fixed stake (LAY_FIXED_STAKE per market):
        Liability = stake * (lay_odds - 1)
        Return    = dead-heat formula applied to that liability

    Dead heat formula (vectorised via _lay_return):
        Win (bad):  ((1 - (exp_winners/act_winners) * odds) / (odds-1)) * liability
        Loss (good): liability / (odds - 1)   [= stake, unchanged by dead heat]
    """
    df      = df.copy()
    lay     = df[lay_odds_col].to_numpy(dtype=float)
    actual  = df[target_col].to_numpy(dtype=float)
    model   = df["Normalised_Model_Odds"].to_numpy(dtype=float)
    is_lay  = model > lay

    exp_winners, act_winners = _build_dh_arrays(df, market_name)

    # Fixed liability
    fl          = LAY_FIXED_LIABILITY[market_name]
    liability_fl = np.full(len(df), fl, dtype=float)
    pnl_fl = np.where(
        is_lay,
        _lay_return(lay, actual, None, exp_winners, act_winners, liability_fl),
        0.0
    )

    # Fixed stake
    fs          = LAY_FIXED_STAKE[market_name]
    liability_fs = fs * np.clip(lay - 1, 1e-8, None)
    pnl_fs = np.where(
        is_lay,
        _lay_return(lay, actual, None, exp_winners, act_winners, liability_fs),
        0.0
    )

    df["Lay_Bet"]              = is_lay
    df["Lay_PnL_FixedLiab"]   = pnl_fl
    df["Lay_PnL_FixedStake"]  = pnl_fs
    return df


def lay_summary(df: pd.DataFrame, target_col: str,
                market_name: str) -> dict:
    bets = df[df["Lay_Bet"]]
    if len(bets) == 0:
        return {
            "Lay_NBets": 0, "Lay_NLost": 0, "Lay_HitRate": np.nan,
            "Lay_PnL_FixedLiab": 0.0,  "Lay_ROI_FixedLiab": np.nan,
            "Lay_PnL_FixedStake": 0.0, "Lay_ROI_FixedStake": np.nan,
        }
    n_bets  = len(bets)
    # "won" for a layer = player did NOT finish in the places
    n_lost  = int(bets[target_col].sum())   # times layer lost (player won/placed)
    n_won   = n_bets - n_lost               # times layer won

    fl      = LAY_FIXED_LIABILITY[market_name]
    fs      = LAY_FIXED_STAKE[market_name]
    pnl_fl  = float(bets["Lay_PnL_FixedLiab"].sum())
    pnl_fs  = float(bets["Lay_PnL_FixedStake"].sum())

    # ROI denominators
    # Fixed liability: total amount at risk = n_bets × liability
    # Fixed stake:     total staked = n_bets × stake (stake varies by odds,
    #                  so use mean stake as the denominator)
    total_liability = n_bets * fl
    lay_odds_vals   = bets["Lay_PnL_FixedStake"]   # proxy; use actual stake col if present
    total_staked_fs = float((fs * (bets[LAY_ODDS_COLS[market_name]] - 1)).sum())

    return {
        "Lay_NBets":            n_bets,
        "Lay_NLost":            n_lost,
        "Lay_HitRate":          round(n_won / n_bets, 4),
        "Lay_PnL_FixedLiab":   round(pnl_fl, 2),
        "Lay_ROI_FixedLiab":   round(pnl_fl / total_liability, 4) if total_liability > 0 else np.nan,
        "Lay_PnL_FixedStake":  round(pnl_fs, 2),
        "Lay_ROI_FixedStake":  round(pnl_fs / total_staked_fs, 4) if total_staked_fs > 0 else np.nan,
    }


# ===== MODEL QUALITY METRICS =====

def compute_discrimination(y_true: np.ndarray, y_prob: np.ndarray) -> dict:
    if y_true.sum() == 0:
        return {"AUC": np.nan, "Avg_Precision": np.nan, "TSS": np.nan,
                "Log_Loss": np.nan, "Brier": np.nan}
    y_clip = np.clip(y_prob, 1e-15, 1 - 1e-15)
    return {
        "AUC":           round(roc_auc_score(y_true, y_prob), 4),
        "Avg_Precision": round(average_precision_score(y_true, y_prob), 4),
        "TSS":           round(tss_optimal(y_true, y_prob), 4),
        "Log_Loss":      round(log_loss(y_true, y_clip), 5),
        "Brier":         round(brier_score_loss(y_true, y_prob), 5),
    }


def calibration_bins(y_true: np.ndarray, y_prob: np.ndarray,
                     n_bins: int = 10) -> pd.DataFrame:
    bins = np.linspace(0, 1, n_bins + 1)
    rows = []
    for lo, hi in zip(bins[:-1], bins[1:]):
        mask = (y_prob >= lo) & (y_prob < hi)
        if mask.sum() == 0:
            continue
        rows.append({
            "Bin_Low":        round(lo, 2),
            "Bin_High":       round(hi, 2),
            "N":              int(mask.sum()),
            "Mean_Predicted": round(float(y_prob[mask].mean()), 4),
            "Actual_Rate":    round(float(y_true[mask].mean()), 4),
        })
    return pd.DataFrame(rows)


# ===== LAY ODDS JOIN =====

LAY_JOIN_COLS = list(LAY_ODDS_COLS.values())   # ["Lay_odds", "Lay_top5", "Lay_top10", "Lay_top20"]
LAY_JOIN_ON   = ["eventID", "playerID"]


def join_lay_odds(df: pd.DataFrame, raw_file: Path) -> pd.DataFrame:
    """
    Left-join Betfair Lay odds onto the processed DataFrame from the raw Excel file.

    The Lay columns are stripped during preprocessing due to high missingness
    (they are not model features) but are needed at backtest time for P&L.

    Join keys: eventID + playerID
    Source:    profit_file (e.g. PGA.xlsx / Euro.xlsx) in TOUR_CONFIG

    Rows with no Lay odds in the raw file will have NaN in those columns;
    backtest_window already skips events where all Lay odds are NaN.
    """
    if raw_file is None or not raw_file.exists():
        print(f"  WARNING: Raw file not found for Lay odds join: {raw_file}")
        return df

    try:
        available_join = [c for c in LAY_JOIN_ON if c in df.columns]
        if not available_join:
            print(f"  WARNING: Join keys {LAY_JOIN_ON} not found in processed data — skipping Lay odds join")
            return df

        # Read only the columns we need — much faster than loading the full raw file
        usecols = available_join + LAY_JOIN_COLS
        raw = pd.read_excel(raw_file, usecols=lambda c: c in usecols)

        # Keep only Lay columns that actually exist in the raw file
        present_lay = [c for c in LAY_JOIN_COLS if c in raw.columns]
        missing_lay = [c for c in LAY_JOIN_COLS if c not in raw.columns]
        if missing_lay:
            print(f"  WARNING: Lay columns not found in raw file: {missing_lay}")
        if not present_lay:
            print(f"  WARNING: No Lay columns found in {raw_file.name} — backtesting will be skipped")
            return df

        raw = raw[available_join + present_lay].dropna(subset=available_join)

        # Deduplicate to one row per (eventID, playerID)
        before = len(raw)
        raw = raw.drop_duplicates(subset=available_join, keep="first")
        if len(raw) < before:
            print(f"  [join_lay_odds] Deduplicated: {before:,} → {len(raw):,} rows")

        # Drop any Lay cols already present in df to avoid _x/_y conflicts
        existing = [c for c in present_lay if c in df.columns]
        if existing:
            df = df.drop(columns=existing)

        df = df.merge(raw[available_join + present_lay], on=available_join, how="left")

        n_matched = df[present_lay[0]].notna().sum()
        print(f"  Lay odds joined: {n_matched:,}/{len(df):,} rows have Lay odds "
              f"({100*n_matched/len(df):.1f}%)")

    except Exception as e:
        print(f"  WARNING: Could not join Lay odds from {raw_file.name}: {e}")

    return df


# ===== WALK-FORWARD WINDOW LOGIC =====

def get_windows(df: pd.DataFrame, min_test_year: int = None):
    """
    Build list of (train_start, train_end, test_year) tuples.
    Requires at least TRAINING_YEARS of prior data for each test year.
    """
    years = sorted(df["Date"].dt.year.dropna().unique())
    windows = []
    for test_year in years:
        train_end   = test_year - 1
        train_start = test_year - TRAINING_YEARS
        if train_start < years[0]:
            continue
        if min_test_year and test_year < min_test_year:
            continue
        windows.append((int(train_start), int(train_end), int(test_year)))
    return windows


# ===== WINDOW TRAINING =====

def train_window(tour_key: str, train_df: pd.DataFrame, test_year: int,
                 n_trials: int, prev_window_dir: Path = None) -> dict:
    """
    Train all four betting markets for one walk-forward window.

    Monkey-patches smt.OPTUNA_TRIALS and smt.MODELS_DIR so tuning uses
    the walk-forward trial budget and saves/loads warm params to the
    window-specific cache directory rather than the production MODELS_DIR.

    Warm-start: best params from prev_window_dir are copied into this
    window's directory before training begins.

    Returns dict: market_name → market_package (or empty dict if no markets trained).
    """
    window_dir = WF_MODELS_DIR / f"{tour_key}_{test_year}"
    window_dir.mkdir(parents=True, exist_ok=True)

    if prev_window_dir and prev_window_dir.exists():
        for f in prev_window_dir.glob("*_best_params.pkl"):
            dest = window_dir / f.name
            if not dest.exists():
                shutil.copy(f, dest)

    orig_trials     = smt.OPTUNA_TRIALS
    orig_models_dir = smt.MODELS_DIR
    smt.OPTUNA_TRIALS = n_trials
    smt.MODELS_DIR    = window_dir

    try:
        package = {}
        for market_name, market_config in BETTING_MARKETS.items():
            target_col = market_config["target_col"]

            if target_col in train_df.columns:
                n_pos = int(train_df[target_col].sum())
                if n_pos < MIN_POSITIVES:
                    print(f"    {market_name}: skipped "
                          f"({n_pos} positives < {MIN_POSITIVES} minimum)")
                    continue

            model_vars = get_market_vars(market_config)
            result = smt.train_market(
                market_name, market_config, train_df, tour_key, model_vars
            )
            if result is not None:
                package[market_name] = result
    finally:
        smt.OPTUNA_TRIALS = orig_trials
        smt.MODELS_DIR    = orig_models_dir

    return package


# ===== WINDOW BACKTESTING =====

def backtest_window(tour_key: str, test_df: pd.DataFrame, test_year: int,
                    package: dict):
    """
    Apply all trained markets to every event in test_df using Betfair Lay odds.
    Returns (all_predictions list, event_summaries list).
    """
    all_predictions = []
    event_summaries = []

    # Clean zero/NaN lay odds once across the whole test year, per market.
    # Zero odds corrupt P&L via division by (lay - 1); NaN rows are already
    # excluded by predict_event's dropna but zeros slip through.
    for market_name, lay_col in LAY_ODDS_COLS.items():
        if lay_col in test_df.columns:
            before   = test_df[lay_col].notna().sum()
            test_df[lay_col] = test_df[lay_col].replace(0, np.nan)
            n_zeroed = before - test_df[lay_col].notna().sum()
            if n_zeroed > 0:
                print(f"  {market_name}: zeroed {n_zeroed} invalid lay odds entries")

    events = test_df["eventID"].unique()
    print(f"  Test {test_year}: {len(events)} events  |  {len(test_df):,} player-rows")

    for event_id in sorted(events):
        event_df   = test_df[test_df["eventID"] == event_id].copy()
        event_date = event_df["Date"].iloc[0].strftime("%Y-%m-%d")

        for market_name, market_pkg in package.items():
            market_config = BETTING_MARKETS[market_name]
            target_col    = market_config["target_col"]
            odds_col      = market_config["odds_col"]
            lay_odds_col  = LAY_ODDS_COLS.get(market_name)

            if target_col not in event_df.columns:
                continue

            # Skip if Lay odds column is absent or fully missing for this event
            if lay_odds_col is None or lay_odds_col not in event_df.columns:
                print(f"    {market_name}: no Lay odds column '{lay_odds_col}' — skipping event {event_id}")
                continue
            if event_df[lay_odds_col].isna().all():
                continue

            preds = predict_event(event_df, market_name, market_pkg)
            if preds is None or len(preds) == 0:
                continue

            preds["Actual"] = preds[target_col].astype(int)
            preds = apply_back_strategy(preds, lay_odds_col, target_col, market_name)
            preds = apply_lay_strategy(preds, lay_odds_col, target_col, market_name)

            bs  = back_summary(preds, target_col)
            ls  = lay_summary(preds, target_col, market_name)

            event_summaries.append({
                "Tour":      tour_key,
                "Test_Year": test_year,
                "EventID":   event_id,
                "Date":      event_date,
                "Market":    market_name,
                "FieldSize": len(preds),
                "Positives": int(preds["Actual"].sum()),
                **bs,
                **ls,
            })

            preds["Tour"]      = tour_key
            preds["Test_Year"] = test_year
            preds["EventID"]   = event_id
            preds["Date"]      = event_date
            preds["Market"]    = market_name
            all_predictions.append(preds)

    return all_predictions, event_summaries


# ===== AGGREGATE ACROSS WINDOWS =====

def aggregate_results(all_preds_list: list, event_summaries_list: list,
                      tour_key: str) -> dict:
    """Combine results from all walk-forward windows into summary DataFrames."""
    if not all_preds_list:
        return None

    pred_df  = pd.concat(all_preds_list, ignore_index=True)
    event_df = (
        pd.DataFrame(event_summaries_list)
        .sort_values(["Market", "Test_Year", "Date", "EventID"])
        .reset_index(drop=True)
    )

    for mkt in event_df["Market"].unique():
        mask = event_df["Market"] == mkt
        event_df.loc[mask, "Back_Cumulative_PnL"]          = event_df.loc[mask, "Back_PnL"].cumsum().round(2).to_numpy()
        event_df.loc[mask, "Lay_Cumulative_PnL_FixedLiab"] = event_df.loc[mask, "Lay_PnL_FixedLiab"].cumsum().round(2).to_numpy()
        event_df.loc[mask, "Lay_Cumulative_PnL_FixedStake"]= event_df.loc[mask, "Lay_PnL_FixedStake"].cumsum().round(2).to_numpy()

    season_rows = []
    for test_year in sorted(pred_df["Test_Year"].unique()):
        for market_name, market_config in BETTING_MARKETS.items():
            target_col = market_config["target_col"]
            mdf = pred_df[
                (pred_df["Test_Year"] == test_year) &
                (pred_df["Market"]    == market_name)
            ]
            if len(mdf) == 0 or target_col not in mdf.columns:
                continue
            y_true = mdf["Actual"].values
            y_prob = mdf["Normalised_Probability"].values
            disc = compute_discrimination(y_true, y_prob)
            bs   = back_summary(mdf, target_col)
            ls   = lay_summary(mdf, target_col, market_name)
            season_rows.append({
                "Tour":       tour_key,
                "Test_Year":  test_year,
                "Market":     market_name,
                "N_Events":   int(mdf["EventID"].nunique()),
                "N_Players":  len(mdf),
                "Prevalence": round(float(y_true.mean()), 4),
                **disc, **bs, **ls,
            })

    summary_rows = []
    calib_sheets = {}
    for market_name, market_config in BETTING_MARKETS.items():
        target_col = market_config["target_col"]
        mdf = pred_df[pred_df["Market"] == market_name]
        if len(mdf) == 0 or target_col not in mdf.columns:
            continue
        y_true = mdf["Actual"].values
        y_prob = mdf["Normalised_Probability"].values
        disc = compute_discrimination(y_true, y_prob)
        bs   = back_summary(mdf, target_col)
        ls   = lay_summary(mdf, target_col, market_name)
        summary_rows.append({
            "Tour":          tour_key,
            "Market":        market_name,
            "N_Test_Years":  int(pred_df[pred_df["Market"] == market_name]["Test_Year"].nunique()),
            "N_Events":      int(mdf["EventID"].nunique()),
            "N_Players":     len(mdf),
            "Prevalence":    round(float(y_true.mean()), 4),
            **disc, **bs, **ls,
        })
        calib_sheets[market_name] = calibration_bins(y_true, y_prob)

    return {
        "summary":         pd.DataFrame(summary_rows),
        "season_summary":  pd.DataFrame(season_rows),
        "event_results":   event_df,
        "all_predictions": pred_df,
        "calibration":     calib_sheets,
    }


# ===== GRID SEARCH =====

def _grid_metrics(band_df: pd.DataFrame, pnl_col: str,
                  actuals_col: str, odds_col: str,
                  staked_per_bet: float | None = None) -> dict:
    """
    Compute standard grid metrics for a filtered DataFrame.

    staked_per_bet: fixed amount staked per bet for ROI denominator.
                    If None, ROI is not computed (e.g. fixed-liability lay).
    """
    pnl_vals = band_df[pnl_col].to_numpy(dtype=float)
    actuals  = band_df[actuals_col].to_numpy(dtype=float)
    odds     = pd.to_numeric(band_df[odds_col], errors="coerce").to_numpy(dtype=float)
    n_bets   = len(band_df)
    total_pnl = pnl_vals.sum()

    event_pnl = pd.Series(pnl_vals, index=band_df.index).groupby(band_df["EventID"]).sum()
    n_events  = len(event_pnl)
    epnl_std  = event_pnl.std()
    sharpe    = (event_pnl.mean() / epnl_std) if epnl_std > 0 else 0
    cum       = event_pnl.cumsum().values
    max_dd    = float((cum - np.maximum.accumulate(cum)).min())

    roi = (total_pnl / (n_bets * staked_per_bet)) if staked_per_bet else np.nan

    return {
        "N_Bets":        n_bets,
        "N_Won":         int(actuals.sum()),
        "Strike_Rate_%": round(actuals.mean() * 100, 2),
        "Total_PnL":     round(total_pnl, 2),
        "ROI_%":         round(roi * 100, 2) if not np.isnan(roi) else np.nan,
        "Avg_Lay_Odds":  round(float(np.nanmean(odds)), 2) if len(odds) > 0 else np.nan,
        "Sharpe":        round(sharpe, 3),
        "Max_Drawdown":  round(max_dd, 2),
    }


def run_back_grid_search(pred_df: pd.DataFrame) -> pd.DataFrame:
    """
    Back grid: sweep rating floor thresholds across edge-filtered predictions.

    Baseline edge condition (model odds < lay odds) always applied first.
    Floors: bet if rating >= threshold. "All" row = no rating filter.
    Results sorted by Market then Rating_Floor.
    """
    df = pred_df.copy()

    df["_Lay_Odds"] = np.nan
    for market_name, lay_col in LAY_ODDS_COLS.items():
        if lay_col in df.columns:
            df.loc[df["Market"] == market_name, "_Lay_Odds"] = df.loc[df["Market"] == market_name, lay_col]

    # Baseline: model odds < lay odds
    df = df[df["Normalised_Model_Odds"] < df["_Lay_Odds"]].copy()

    rf  = df["DeadHeat_RF"].to_numpy(dtype=float) if "DeadHeat_RF" in df.columns else np.ones(len(df))
    lay = df["_Lay_Odds"].to_numpy(dtype=float)
    act = df["Actual"].to_numpy(dtype=float)
    df["_Back_PnL"] = np.where(act == 1, BACK_STAKE * (rf * lay - 1), -BACK_STAKE)

    thresholds = [("All", None)] + [(f">= {t}", t) for t in GRID_BACK_RATING_FLOORS]

    print(f"\n  Back grid search: {len(BETTING_MARKETS)} markets × {len(thresholds)} "
          f"rating floors (edge-filtered: {len(df):,} bets)...")

    results = []
    for market_name in BETTING_MARKETS:
        mdf = df[df["Market"] == market_name]
        if len(mdf) == 0:
            continue
        for label, floor in thresholds:
            band_df = mdf[mdf["rating"] >= floor] if floor is not None else mdf
            if len(band_df) == 0:
                continue
            m = _grid_metrics(band_df, "_Back_PnL", "Actual", "_Lay_Odds", BACK_STAKE)
            results.append({
                "Market":        market_name,
                "Rating_Floor":  floor if floor is not None else "All",
                **m,
            })

    if not results:
        return pd.DataFrame()

    grid_df = (pd.DataFrame(results)
               .sort_values(["Market", "Rating_Floor"], na_position="first")
               .reset_index(drop=True))

    print(f"  Back grid complete: {len(grid_df)} rows")
    for _, row in grid_df.iterrows():
        print(f"    {row['Market']:<8}  Rating>={str(row['Rating_Floor']):<6}  "
              f"{row['N_Bets']:>5} bets  "
              f"P&L=£{row['Total_PnL']:>8.2f}  ROI={row['ROI_%']:>6.1f}%")

    return grid_df


def run_lay_grid_search(pred_df: pd.DataFrame) -> pd.DataFrame:
    """
    Lay grid: sweep rating ceiling thresholds across edge-filtered predictions.

    Baseline edge condition (model odds > lay odds) always applied first.
    Ceilings: bet if rating < threshold. "All" row = no rating filter.

    Both fixed-liability and fixed-stake P&L columns are reported.
    Results sorted by Market then Rating_Ceiling.
    """
    df = pred_df.copy()

    df["_Lay_Odds"] = np.nan
    for market_name, lay_col in LAY_ODDS_COLS.items():
        if lay_col in df.columns:
            df.loc[df["Market"] == market_name, "_Lay_Odds"] = df.loc[df["Market"] == market_name, lay_col]

    # Baseline: model odds > lay odds
    df = df[df["Normalised_Model_Odds"] > df["_Lay_Odds"]].copy()

    thresholds = [("All", None)] + [(f"< {t}", t) for t in GRID_LAY_RATING_CEILINGS]

    print(f"\n  Lay grid search: {len(BETTING_MARKETS)} markets × {len(thresholds)} "
          f"rating ceilings (edge-filtered: {len(df):,} bets)...")

    results = []
    for market_name, market_config in BETTING_MARKETS.items():
        target_col = market_config["target_col"]
        mdf = df[df["Market"] == market_name]
        if len(mdf) == 0:
            continue

        fl = LAY_FIXED_LIABILITY[market_name]
        fs = LAY_FIXED_STAKE[market_name]

        for label, ceiling in thresholds:
            band_df = mdf[mdf["rating"] < ceiling] if ceiling is not None else mdf
            if len(band_df) == 0:
                continue

            # Fixed liability metrics — ROI denominator = n_bets × liability
            m_fl = _grid_metrics(band_df, "Lay_PnL_FixedLiab",
                                 "Actual", "_Lay_Odds", fl)
            # Fixed stake metrics — ROI denominator = n_bets × stake
            m_fs = _grid_metrics(band_df, "Lay_PnL_FixedStake",
                                 "Actual", "_Lay_Odds", fs)

            results.append({
                "Market":                  market_name,
                "Rating_Ceiling":          ceiling if ceiling is not None else "All",
                "N_Bets":                  m_fl["N_Bets"],
                "N_Lost":                  m_fl["N_Won"],    # N_Won = player placed = layer lost
                "Strike_Rate_%":           round((m_fl["N_Bets"] - m_fl["N_Won"]) / m_fl["N_Bets"] * 100, 2),
                "Avg_Lay_Odds":            m_fl["Avg_Lay_Odds"],
                # Fixed liability
                "FL_Liability":            fl,
                "FL_Total_PnL":            m_fl["Total_PnL"],
                "FL_ROI_%":                m_fl["ROI_%"],
                "FL_Sharpe":               m_fl["Sharpe"],
                "FL_Max_Drawdown":         m_fl["Max_Drawdown"],
                # Fixed stake
                "FS_Stake":                fs,
                "FS_Total_PnL":            m_fs["Total_PnL"],
                "FS_ROI_%":                m_fs["ROI_%"],
                "FS_Sharpe":               m_fs["Sharpe"],
                "FS_Max_Drawdown":         m_fs["Max_Drawdown"],
            })

    if not results:
        return pd.DataFrame()

    grid_df = (pd.DataFrame(results)
               .sort_values(["Market", "Rating_Ceiling"], na_position="first")
               .reset_index(drop=True))

    print(f"  Lay grid complete: {len(grid_df)} rows")
    for _, row in grid_df.iterrows():
        print(f"    {row['Market']:<8}  Rating<{str(row['Rating_Ceiling']):<6}  "
              f"{row['N_Bets']:>5} bets  "
              f"FL P&L=£{row['FL_Total_PnL']:>8.2f}  "
              f"FS P&L=£{row['FS_Total_PnL']:>8.2f}")

    return grid_df


def _write_grid_sheet(wb, grid_df: pd.DataFrame, sheet_name: str,
                      pnl_cols: list[str], roi_cols: list[str]):
    """
    Write a grid search DataFrame to a styled Excel sheet.

    pnl_cols: column names to colour-code by P&L value
    roi_cols: column names to colour-code by ROI sign
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    thin_side  = Side(style="thin", color="CCCCCC")
    std_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center     = Alignment(horizontal="center", vertical="center")
    HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HDR_FILL   = PatternFill("solid", start_color="1F4E79")
    BODY_FONT  = Font(name="Arial", size=9)
    POS_FILL   = PatternFill("solid", start_color="C6EFCE")
    NEG_FILL   = PatternFill("solid", start_color="FFC7CE")
    MID_FILL   = PatternFill("solid", start_color="FFEB9C")

    cols = list(grid_df.columns)
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = center
        cell.border    = std_border

    pnl_idxs = [cols.index(c) + 1 for c in pnl_cols if c in cols]
    roi_idxs = [cols.index(c) + 1 for c in roi_cols if c in cols]

    for row_idx, row in grid_df.iterrows():
        excel_row = row_idx + 2
        for col_idx, val in enumerate(row.values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.border    = std_border
            cell.alignment = center
        for ci in pnl_idxs:
            pnl = row.iloc[ci - 1]
            ws.cell(row=excel_row, column=ci).fill = (
                POS_FILL if pnl > 0 else NEG_FILL if pnl < -200 else MID_FILL
            )
        for ci in roi_idxs:
            ws.cell(row=excel_row, column=ci).fill = (
                POS_FILL if row.iloc[ci - 1] > 0 else NEG_FILL
            )

    default_w = 12
    col_widths = {
        "Market": 10, "Rating_Floor": 14, "Rating_Ceiling": 16,
        "N_Bets": 9, "N_Won": 8, "N_Lost": 8, "Strike_Rate_%": 14,
        "Total_PnL": 13, "ROI_%": 9, "Avg_Lay_Odds": 13,
        "Sharpe": 10, "Max_Drawdown": 15,
        "FL_Liability": 13, "FL_Total_PnL": 14, "FL_ROI_%": 10,
        "FL_Sharpe": 10, "FL_Max_Drawdown": 16,
        "FS_Stake": 10, "FS_Total_PnL": 14, "FS_ROI_%": 10,
        "FS_Sharpe": 10, "FS_Max_Drawdown": 16,
    }
    for col_idx, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, default_w)

    ws.freeze_panes = "A2"


# ===== EXPORT =====

def export_results(results: dict, tour_key: str,
                   back_grid_df: pd.DataFrame = None,
                   lay_grid_df: pd.DataFrame = None) -> Path:
    WF_RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = WF_RESULTS_DIR / f"{tour_key}_WalkForward_Backtest.xlsx"

    id_cols   = [c for c in ["Test_Year", "Date", "EventID", "Market",
                              "playerID", "surname", "firstname", "posn", "rating"]
                 if c in results["all_predictions"].columns]
    lay_cols  = [c for c in LAY_ODDS_COLS.values()
                 if c in results["all_predictions"].columns]
    pred_cols = ["Model_Score", "Probability", "Normalised_Probability",
                 "Normalised_Model_Odds", "Actual"]
    bet_cols  = ["Back_Bet", "DeadHeat_RF", "Back_PnL",
                 "Lay_Bet", "Lay_PnL_FixedLiab", "Lay_PnL_FixedStake"]
    all_cols  = id_cols + lay_cols + pred_cols + bet_cols
    export_df = results["all_predictions"][
        [c for c in all_cols if c in results["all_predictions"].columns]
    ]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        results["summary"].to_excel(        writer, sheet_name="Summary",         index=False)
        results["season_summary"].to_excel( writer, sheet_name="Season_Summary",  index=False)
        results["event_results"].to_excel(  writer, sheet_name="Event_Results",   index=False)
        export_df.to_excel(                 writer, sheet_name="All_Predictions", index=False)

    wb = load_workbook(out_path)
    if back_grid_df is not None and len(back_grid_df) > 0:
        _write_grid_sheet(wb, back_grid_df,
                          sheet_name="Back_Strategy_Grid",
                          pnl_cols=["Total_PnL"],
                          roi_cols=["ROI_%"])
    if lay_grid_df is not None and len(lay_grid_df) > 0:
        _write_grid_sheet(wb, lay_grid_df,
                          sheet_name="Lay_Strategy_Grid",
                          pnl_cols=["FL_Total_PnL", "FS_Total_PnL"],
                          roi_cols=["FL_ROI_%", "FS_ROI_%"])
    wb.save(out_path)

    print(f"\n  Saved: {out_path}")
    return out_path


# ===== TOUR RUNNER =====

def run_tour(tour_key: str, tour_info: dict, n_trials: int,
             start_year: int, min_test_year: int, force_retrain: bool):
    print(f"\n{'='*60}")
    print(f"WALK-FORWARD: {tour_info['name']}")
    print(f"{'='*60}")

    hist_path = tour_info["historical_file"]
    if not hist_path.exists():
        print(f"  Processed file not found: {hist_path}")
        return None

    df = pd.read_excel(hist_path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.dropna(subset=["Date"])

    # Trim to start_year if specified — reduces memory and speeds up window generation
    if start_year is not None:
        before = len(df)
        df = df[df["Date"].dt.year >= start_year].copy()
        print(f"  --start-year {start_year}: trimmed {before:,} → {len(df):,} rows")

    # Join Betfair Lay odds from raw file (stripped during preprocessing)
    df = join_lay_odds(df, tour_info.get("profit_file"))

    print(f"  Loaded {len(df):,} rows  |  years: "
          f"{int(df['Date'].dt.year.min())}–{int(df['Date'].dt.year.max())}")

    windows = get_windows(df, min_test_year=min_test_year)
    if not windows:
        print("  No valid walk-forward windows found."); return None

    print(f"\n  Walk-forward plan ({len(windows)} windows):")
    for ts, te, ty in windows:
        cached = (WF_MODELS_DIR / f"{tour_key}_{ty}" / "market_bundle.pkl").exists()
        status = " [cached]" if cached and not force_retrain else ""
        print(f"    Train {ts}–{te}  →  Test {ty}{status}")

    all_preds_list  = []
    event_summ_list = []
    prev_window_dir = None

    for train_start, train_end, test_year in windows:
        print(f"\n  --- Train {train_start}–{train_end} → Test {test_year} ---")

        bundle_path = WF_MODELS_DIR / f"{tour_key}_{test_year}" / "market_bundle.pkl"

        if bundle_path.exists() and not force_retrain:
            print(f"  Loading cached bundle for {test_year}...")
            package = joblib.load(bundle_path)
        else:
            train_df = df[
                (df["Date"].dt.year >= train_start) &
                (df["Date"].dt.year <= train_end)
            ].copy()
            print(f"  Training rows: {len(train_df):,}")

            package = train_window(
                tour_key, train_df, test_year, n_trials, prev_window_dir
            )
            if not package:
                print(f"  No markets trained — skipping test year {test_year}")
                prev_window_dir = WF_MODELS_DIR / f"{tour_key}_{test_year}"
                continue

            bundle_path.parent.mkdir(parents=True, exist_ok=True)
            joblib.dump(package, bundle_path)
            print(f"  Bundle saved: {bundle_path.parent.name}/market_bundle.pkl")

        prev_window_dir = WF_MODELS_DIR / f"{tour_key}_{test_year}"

        test_df = df[df["Date"].dt.year == test_year].copy()
        if len(test_df) == 0:
            print(f"  No test data for {test_year}"); continue

        window_preds, window_events = backtest_window(
            tour_key, test_df, test_year, package
        )
        all_preds_list.extend(window_preds)
        event_summ_list.extend(window_events)

    results = aggregate_results(all_preds_list, event_summ_list, tour_key)
    if results is None:
        print("  No results to aggregate."); return None

    print(f"\n  === AGGREGATE WALK-FORWARD RESULTS ({len(windows)} windows) ===")
    print(f"  {'Market':<8}  {'AUC':>6}  {'AP':>6}  {'TSS':>6}  {'Back_PnL':>10}  {'Back_ROI':>9}")
    for _, row in results["summary"].iterrows():
        print(
            f"  {row['Market']:<8}  {row['AUC']:>6.4f}  {row['Avg_Precision']:>6.4f}  "
            f"{row['TSS']:>6.4f}  £{row['Back_PnL']:>9.2f}  {row['Back_ROI']:>8.1%}"
        )

    grid_back_df = run_back_grid_search(results["all_predictions"])
    grid_lay_df  = run_lay_grid_search(results["all_predictions"])

    return results, grid_back_df, grid_lay_df


# ===== MAIN =====

def main():
    args = parse_args()

    print("=== GOLF MODEL WALK-FORWARD BACKTESTING ===")
    print(f"Training window: {TRAINING_YEARS} years  |  Optuna trials: {args.trials}")
    print(f"Back stake: £{BACK_STAKE}  |  Lay odds P&L  |  Dead-heat RF applied to place markets")
    if args.start_year:
        print(f"  Data start year: {args.start_year}")
    if args.min_year:
        print(f"  Min test year:   {args.min_year}")
    if args.force_retrain:
        print("  ** Force retrain: ignoring all cached models **")

    tours_to_run = {
        k: v for k, v in TOUR_CONFIG.items()
        if args.tour is None or k == args.tour
    }

    for tour_key, tour_info in tours_to_run.items():
        outcome = run_tour(
            tour_key, tour_info,
            n_trials=args.trials,
            start_year=args.start_year,
            min_test_year=args.min_year,
            force_retrain=args.force_retrain,
        )
        if outcome:
            results, grid_back_df, grid_lay_df = outcome
            export_results(results, tour_key, grid_back_df, grid_lay_df)


if __name__ == "__main__":
    main()