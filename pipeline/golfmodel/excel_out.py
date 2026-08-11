"""One shared Excel writer: bold header, frozen top row, sensible widths."""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_FILL = PatternFill("solid", start_color="1F4E79")


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    """Write {sheet_name: DataFrame} to one styled workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.sheets[name[:31]]
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = _HDR_FONT
                cell.fill = _HDR_FILL
                width = max(10, min(24, len(str(col_name)) + 4))
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.freeze_panes = "A2"
    return path
